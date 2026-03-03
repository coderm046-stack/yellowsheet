import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Consolidated Marksheet Pro", layout="wide")
st.title("🏫 Student Exam Data Consolidator")

def custom_round(x):
    try:
        val = float(x)
        return int(np.floor(val + 0.5))
    except:
        return 0

def clean_marks(val):
    if isinstance(val, str):
        v = val.strip().upper()
        if v == 'AB' or v == '':
            return 0.0
    try:
        return float(val)
    except:
        return 0.0

uploaded_file = st.file_uploader("Upload Excel Marksheet", type="xlsx")

if uploaded_file:
    try:
        xl = pd.ExcelFile(uploaded_file)
        subj_list = [f'Sub{i+1}' for i in range(6)]
        fetch_subj_map = {f'SUB{i+1}': f'Sub{i+1}' for i in range(6)}

        exam_configs = [
            {'label': 'FIRST UNIT TEST (25)',  'sheets': ['FIRST UNIT TEST']},
            {'label': 'FIRST TERM EXAM (50)',  'sheets': ['FIRST TERM']},
            {'label': 'SECOND UNIT TEST (25)', 'sheets': ['SECOND UNIT TEST']},
            {'label': 'ANNUAL EXAM (70/80)',   'sheets': ['ANNUAL EXAM']}
        ]

        result_cols = ['Grand Total', '%', 'Result', 'Remark', 'Rank']
        all_students = {}

        for config in exam_configs:
            sheet_name = next((s for s in xl.sheet_names if s.strip().upper() in config['sheets']), None)
            if sheet_name:
                df = xl.parse(sheet_name)
                df.columns = df.columns.astype(str).str.strip().str.upper()
                t_col = next((c for c in df.columns if 'TOTAL' in c), None)
                p_col = next((c for c in df.columns if '%' in c or 'PERCENT' in c), None)
                r_col = next((c for c in df.columns if 'RESULT' in c), None)

                for _, row in df.iterrows():
                    roll = str(row.get('ROLL NO.', '')).strip()
                    if not roll or roll == 'nan':
                        continue
                    if roll not in all_students:
                        all_students[roll] = {'Name': row.get('STUDENT NAME', 'Unknown'), 'Exams': {}}
                    marks = {v: (str(row.get(k, 0)).strip() if str(row.get(k, 0)).strip().upper() == 'AB'
                                 else row.get(k, 0)) for k, v in fetch_subj_map.items()}
                    marks['Grand Total'] = str(row.get(t_col, '')) if t_col else ''
                    try:
                        raw_p = row.get(p_col, '')
                        marks['%'] = str(round(float(raw_p), 2)) if raw_p != '' else ''
                    except:
                        marks['%'] = str(row.get(p_col, ''))
                    marks['Result'] = str(row.get(r_col, ''))
                    all_students[roll]['Exams'][config['label']] = marks

        categories = [
            'FIRST UNIT TEST (25)', 'FIRST TERM EXAM (50)', 'SECOND UNIT TEST (25)',
            'ANNUAL EXAM (70/80)', 'INT/PRACTICAL (20/30)',
            'Total Marks Out of 200', 'Average Marks 200/2=100'
        ]

        student_rolls = sorted(
            all_students.keys(),
            key=lambda x: float(x) if x.replace('.', '', 1).isdigit() else 0
        )

        # Build base dataframe
        rows = []
        for roll in student_rolls:
            s = all_students[roll]
            for cat in categories:
                row_data = {
                    'Roll No.': roll if cat == 'FIRST UNIT TEST (25)' else '',
                    'Column1':  s['Name'] if cat == 'FIRST UNIT TEST (25)' else '',
                    'Column2':  cat
                }
                for col in subj_list + result_cols:
                    row_data[col] = ''
                if cat in s['Exams']:
                    row_data.update(s['Exams'][cat])
                elif cat == 'INT/PRACTICAL (20/30)':
                    for sub in subj_list:
                        row_data[sub] = "0"
                rows.append(row_data)

        base_df = pd.DataFrame(rows)
        for col in base_df.columns:
            base_df[col] = base_df[col].astype(str).replace('nan', '')

        # ── Internal Marks Input ────────────────────────────────────────────────────
        st.markdown("---")
        st.subheader("📝 Enter Internal / Practical Marks (20/30)")
        st.info("Fill in the internal/practical marks for each student. These are used to compute totals, percentage, result and rank.")

        if 'internal_marks' not in st.session_state:
            st.session_state.internal_marks = {
                roll: {sub: "0" for sub in subj_list}
                for roll in student_rolls
            }

        hdr = st.columns([1, 2] + [1]*6)
        hdr[0].markdown("**Roll No.**")
        hdr[1].markdown("**Student Name**")
        for i, sub in enumerate(subj_list):
            hdr[i+2].markdown(f"**{sub}**")

        for roll in student_rolls:
            name = all_students[roll]['Name']
            cols = st.columns([1, 2] + [1]*6)
            cols[0].write(roll)
            cols[1].write(name)
            for i, sub in enumerate(subj_list):
                val = cols[i+2].text_input(
                    label=f"{roll}-{sub}",
                    value=st.session_state.internal_marks[roll].get(sub, "0"),
                    key=f"int_{roll}_{sub}",
                    label_visibility="collapsed"
                )
                st.session_state.internal_marks[roll][sub] = val

        # Inject internal marks into base_df
        for i, roll in enumerate(student_rolls):
            int_row_idx = i * 7 + 4
            for sub in subj_list:
                base_df.at[int_row_idx, sub] = st.session_state.internal_marks[roll][sub]

        st.markdown("---")
        st.subheader("📊 Marks Preview & Edit")
        edited_df = st.data_editor(base_df, hide_index=True, use_container_width=True)

        # ── Generate Report ─────────────────────────────────────────────────────────
        if st.button("🚀 Generate Final Report & Rank"):

            # Step 1: compute per-student results
            student_results = []
            for s_idx, roll in enumerate(student_rolls):
                block = edited_df.iloc[s_idx*7 : s_idx*7+7].copy()
                nums  = block.iloc[0:5][subj_list].applymap(clean_marks)
                t200  = nums.sum()
                a100  = t200.apply(lambda x: custom_round(x / 2))
                gt    = int(a100.sum())
                pc    = round((gt / 600) * 100, 2)
                isp   = all(m >= 35 for m in a100)
                student_results.append({
                    'roll': roll,
                    'name': all_students[roll]['Name'],
                    't200': t200,
                    'a100': a100,
                    'gt':   gt,
                    'pc':   pc,
                    'pass': isp
                })

            # Step 2: compute dense rank (only PASS students)
            pass_totals_sorted = sorted(
                set(sr['gt'] for sr in student_results if sr['pass']),
                reverse=True
            )
            rank_map = {gt_val: rank+1 for rank, gt_val in enumerate(pass_totals_sorted)}
            for sr in student_results:
                sr['rank'] = rank_map[sr['gt']] if sr['pass'] else ''

            # Step 3: rebuild final_df
            processed = []
            for s_idx, sr in enumerate(student_results):
                block = edited_df.iloc[s_idx*7 : s_idx*7+7].copy().reset_index(drop=True)
                t200, a100 = sr['t200'], sr['a100']
                for sub in subj_list:
                    block.at[5, sub] = str(int(t200[sub]))
                    block.at[6, sub] = str(int(a100[sub]))
                block.at[6, 'Grand Total'] = str(sr['gt'])
                block.at[6, '%']           = str(sr['pc'])
                block.at[6, 'Result']      = "PASS" if sr['pass'] else "FAIL"
                block.at[6, 'Rank']        = str(sr['rank'])
                processed.append(block)

            final_df = pd.concat(processed).reset_index(drop=True)

            # Show summary
            st.success(f"✅ Report Generated! {sum(1 for sr in student_results if sr['pass'])} students PASSED.")
            summary_df = pd.DataFrame([{
                'Roll No.':    sr['roll'],
                'Name':        sr['name'],
                'Grand Total': sr['gt'],
                '%':           sr['pc'],
                'Result':      'PASS' if sr['pass'] else 'FAIL',
                'Rank':        sr['rank']
            } for sr in student_results])
            st.subheader("📋 Result Summary")
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

            # ── Build Excel with live formulas ─────────────────────────────────────
            wb = Workbook()
            ws = wb.active
            ws.title = "Consolidated"

            # Hidden helper sheet: col A = GT values, col B = IsPass (1/0)
            ws_h = wb.create_sheet("_RankHelper")
            ws_h.sheet_state = "hidden"
            ws_h.cell(row=1, column=1, value="GT")
            ws_h.cell(row=1, column=2, value="IsPass")

            # Styles
            hdr_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            hdr_fill  = PatternFill("solid", start_color="1F4E79")
            cat_fill  = {
                'FIRST UNIT TEST (25)':    PatternFill("solid", start_color="DDEBF7"),
                'FIRST TERM EXAM (50)':    PatternFill("solid", start_color="E2EFDA"),
                'SECOND UNIT TEST (25)':   PatternFill("solid", start_color="FFF2CC"),
                'ANNUAL EXAM (70/80)':     PatternFill("solid", start_color="FCE4D6"),
                'INT/PRACTICAL (20/30)':   PatternFill("solid", start_color="EAD1DC"),
                'Total Marks Out of 200':  PatternFill("solid", start_color="D9D9D9"),
                'Average Marks 200/2=100': PatternFill("solid", start_color="BDD7EE"),
            }
            thin   = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'),  bottom=Side(style='thin'))
            ctr    = Alignment(horizontal='center', vertical='center')

            all_cols = ['Roll No.', 'Column1', 'Column2'] + subj_list + result_cols
            hdrs     = ['Roll No.', 'Student Name', 'Exam Type'] + subj_list + result_cols
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font=hdr_font; c.fill=hdr_fill; c.alignment=ctr; c.border=thin

            widths = {'Roll No.':10,'Column1':22,'Column2':28}
            for sub in subj_list:  widths[sub] = 10
            for rc  in result_cols: widths[rc]  = 13
            for ci, cn in enumerate(all_cols, 1):
                ws.column_dimensions[get_column_letter(ci)].width = widths.get(cn, 12)

            SUB_S  = 4              # column D  (Sub1)
            GT_C   = SUB_S + 6     # column J  (Grand Total)
            PCT_C  = GT_C + 1      # column K  (%)
            RES_C  = PCT_C + 1     # column L  (Result)
            REM_C  = RES_C + 1     # column M  (Remark)
            RNK_C  = REM_C + 1     # column N  (Rank)

            sub_lets = [get_column_letter(SUB_S + i) for i in range(6)]
            gt_let   = get_column_letter(GT_C)
            res_let  = get_column_letter(RES_C)

            n        = len(student_rolls)
            h_gt_rng = f"_RankHelper!$A$2:$A${n+1}"   # helper GT range

            avg_excel_rows = []   # track excel row of each student's avg row

            for s_idx, roll in enumerate(student_rolls):
                sr   = student_results[s_idx]
                brow = 2 + s_idx * 7   # first data row in Excel for this student

                for cat_idx, cat in enumerate(categories):
                    erow = brow + cat_idx
                    fl   = cat_fill.get(cat, PatternFill("solid", start_color="FFFFFF"))

                    ws.cell(row=erow, column=1, value=roll if cat_idx==0 else "")
                    ws.cell(row=erow, column=2, value=sr['name'] if cat_idx==0 else "")
                    ws.cell(row=erow, column=3, value=cat)

                    if cat == 'Total Marks Out of 200':
                        r1, r5 = brow, brow+4
                        for i, sl in enumerate(sub_lets):
                            c = ws.cell(row=erow, column=SUB_S+i,
                                        value=f"=SUM({sl}{r1}:{sl}{r5})")
                            c.fill=fl; c.border=thin; c.alignment=ctr
                            c.font=Font(name="Arial", bold=True)
                        for ri in range(len(result_cols)):
                            c=ws.cell(row=erow, column=GT_C+ri, value="")
                            c.fill=fl; c.border=thin

                    elif cat == 'Average Marks 200/2=100':
                        trow = erow - 1
                        for i, sl in enumerate(sub_lets):
                            c = ws.cell(row=erow, column=SUB_S+i,
                                        value=f"=ROUND({sl}{trow}/2,0)")
                            c.fill=fl; c.border=thin; c.alignment=ctr
                            c.font=Font(name="Arial", bold=True)

                        c = ws.cell(row=erow, column=GT_C,
                                    value=f"=SUM({sub_lets[0]}{erow}:{sub_lets[-1]}{erow})")
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        c.font=Font(name="Arial", bold=True, color="1F4E79")

                        c = ws.cell(row=erow, column=PCT_C,
                                    value=f"=ROUND({gt_let}{erow}/600*100,2)")
                        c.fill=fl; c.border=thin; c.alignment=ctr

                        pass_chk = ",".join([f"{sl}{erow}>=35" for sl in sub_lets])
                        c = ws.cell(row=erow, column=RES_C,
                                    value=f'=IF(AND({pass_chk}),"PASS","FAIL")')
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        c.font=Font(name="Arial", bold=True)

                        c=ws.cell(row=erow, column=REM_C, value="")
                        c.fill=fl; c.border=thin

                        # Rank — filled below
                        c=ws.cell(row=erow, column=RNK_C, value="")
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        c.font=Font(name="Arial", bold=True, color="C00000")

                        # Helper sheet: point to this row's GT and Result
                        h_row = s_idx + 2
                        ws_h.cell(row=h_row, column=1,
                                  value=f"=Consolidated!{gt_let}{erow}")
                        ws_h.cell(row=h_row, column=2,
                                  value=f'=IF(Consolidated!{res_let}{erow}="PASS",1,0)')

                        avg_excel_rows.append((erow, h_row))

                    else:
                        frow = final_df.iloc[s_idx*7 + cat_idx]
                        for i, sub in enumerate(subj_list):
                            v = frow.get(sub, "")
                            try: v = float(v)
                            except: pass
                            c = ws.cell(row=erow, column=SUB_S+i, value=v)
                            c.fill=fl; c.border=thin; c.alignment=ctr

                        for ri, rc in enumerate(result_cols):
                            v = "" if rc == 'Rank' else frow.get(rc, "")
                            c = ws.cell(row=erow, column=GT_C+ri, value=v)
                            c.fill=fl; c.border=thin; c.alignment=ctr

                    for ci in [1, 2, 3]:
                        c = ws.cell(row=erow, column=ci)
                        c.fill=fl; c.border=thin
                        c.font=Font(name="Arial", bold=(ci==2 and cat_idx==0))

            # ── RANK formulas ──────────────────────────────────────────────────────
            # For each student:
            #   IF IsPass=1,  COUNTIF(all_GT_range, ">" & this_GT) + 1,  ""
            # This gives rank 1 to highest scorer, handles ties correctly.
            # FAIL students get blank.
            for (erow, h_row) in avg_excel_rows:
                rank_formula = (
                    f'=IF(_RankHelper!$B${h_row}=1,'
                    f'COUNTIF({h_gt_rng},">"&_RankHelper!$A${h_row})+1,"")'
                )
                fl = cat_fill['Average Marks 200/2=100']
                c  = ws.cell(row=erow, column=RNK_C, value=rank_formula)
                c.fill=fl; c.border=thin; c.alignment=ctr
                c.font=Font(name="Arial", bold=True, color="C00000")

            ws.freeze_panes = "A2"

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                "📥 Download Excel (with Live Formulas)",
                output.getvalue(),
                "Final_Consolidated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error: {e}")
        import traceback
        st.code(traceback.format_exc())
